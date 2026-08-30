"""Importa el export propio de la app FitDays.

Es la unica via para las metricas que Health Connect no contempla: grasa
visceral y subcutanea, musculo esqueletico, proteina y edad corporal.

Donde esta el export, que la app no documenta:
    Tablas -> History Records -> una fecha -> boton "+" arriba -> Exportar

Rarezas del fichero, verificadas contra el formato que documento
https://gitlab.com/leroyzwakman1/runalyze-fitdays :

  * la extension dice .csv pero suele ser un Excel antiguo (OLE2/CDFV2);
    tambien se ha visto .xlsx y CSV de verdad, asi que se olfatea el tipo
  * los valores llevan la unidad pegada: "60.5kg", "13.6%", "1437.0kcal"
  * los huecos son "- -", no celdas vacias
  * la fecha va como "15:35 22/01/2022": hora primero y dia antes que mes
  * las cabeceras estan en el idioma de la app
"""

from __future__ import annotations

import csv
import io
import logging
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from dateutil import parser as dateparser

from ..models import BodyMeasurement

log = logging.getLogger(__name__)

# campo del modelo -> nombres de columna en los idiomas que hemos visto
COLUMNS: dict[str, tuple[str, ...]] = {
    "measured_at":              ("fecha", "date", "datum", "time", "hora"),
    "weight_kg":                ("peso", "weight", "gewicht"),
    "bmi":                      ("imc", "bmi"),
    "fat_percent":              ("grasa corporal", "body fat", "lichaamsvet", "grasa"),
    "subcutaneous_fat_percent": ("grasa subcutanea", "subcutaneous fat", "onderhuids vet"),
    "heart_rate":               ("frecuencia cardiaca", "heart rate", "hartslag", "pulso"),
    "visceral_fat":             ("grasa visceral", "visceral fat", "visceraal vet"),
    "water_percent":            ("agua corporal", "agua", "body water", "watergewicht"),
    "skeletal_muscle_percent":  ("musculo esqueletico", "skeletal muscle", "skeletspier"),
    "muscle_mass_kg":           ("masa muscular", "muscle mass", "spiermassa"),
    "bone_mass_kg":             ("masa osea", "masa esqueletica", "bone mass", "botmassa"),
    "protein_percent":          ("proteina", "protein", "eiwit"),
    "bmr_kcal":                 ("bmr", "metabolismo basal", "basal"),
    "body_age":                 ("edad corporal", "body age", "lichaamsleeftijd"),
}
# el orden documentado, por si las cabeceras no se reconocen
POSITIONAL = ["measured_at", "weight_kg", "bmi", "fat_percent",
              "subcutaneous_fat_percent", "heart_rate", None, "visceral_fat",
              "water_percent", "skeletal_muscle_percent", "muscle_mass_kg",
              "bone_mass_kg", "protein_percent", "bmr_kcal", "body_age"]

NULLS = {"", "-", "- -", "--", "n/a", "null", "0.0kg"}


def _norm(text: str) -> str:
    t = unicodedata.normalize("NFKD", str(text).strip().lower())
    return "".join(c for c in t if not unicodedata.combining(c))


def _num(value) -> float | None:
    """Quita la unidad pegada: "60.5kg" -> 60.5, "13.6%" -> 13.6."""
    if value is None:
        return None
    s = _norm(value)
    if s in NULLS:
        return None
    m = re.search(r"-?\d+(?:[.,]\d+)?", s)
    return float(m.group().replace(",", ".")) if m else None


def _when(value) -> datetime | None:
    if value is None:
        return None
    s = str(value).strip()
    if _norm(s) in NULLS:
        return None
    # "15:35 22/01/2022" -> la hora va delante, hay que reordenar
    m = re.match(r"^(\d{1,2}:\d{2}(?::\d{2})?)\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})$", s)
    if m:
        s = f"{m.group(2)} {m.group(1)}"
    try:
        return dateparser.parse(s, dayfirst=True)
    except (ValueError, OverflowError):
        log.warning("Fecha no reconocida: %r", value)
        return None


def _rows(path: Path) -> list[list]:
    """Devuelve la tabla, olfateando si es OLE2, xlsx o texto."""
    head = path.open("rb").read(8)

    if head.startswith(b"\xd0\xcf\x11\xe0"):          # OLE2: Excel antiguo
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "El fichero es un Excel antiguo (OLE2). Instala el lector con:\n"
                "  pip install xlrd\n"
                "o convierte a CSV con:  ssconvert entrada.csv salida.csv"
            ) from exc
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)]

    if head.startswith(b"PK"):                        # xlsx
        from openpyxl import load_workbook
        ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
        return [list(r) for r in ws.iter_rows(values_only=True)]

    text = path.read_text(encoding="utf-8-sig", errors="replace")
    if "<table" in text[:2000].lower():               # a veces es HTML
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", text, re.S | re.I)
        rows, width = [], text.lower().count("<td") // max(1, text.lower().count("<tr"))
        for i in range(0, len(cells), max(1, width)):
            rows.append([re.sub(r"<[^>]+>", "", c).strip() for c in cells[i:i + width]])
        return rows
    dialect = csv.Sniffer().sniff(text[:2000], delimiters=",;\t")
    return [r for r in csv.reader(io.StringIO(text), dialect)]


def load(path: str | Path) -> list[BodyMeasurement]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    table = [r for r in _rows(path) if any(str(c).strip() for c in r)]
    if len(table) < 2:
        log.warning("El export de FitDays no tiene filas: %s", path)
        return []

    header = [_norm(c) for c in table[0]]
    mapping: dict[int, str] = {}
    for field, names in COLUMNS.items():
        for i, col in enumerate(header):
            if i in mapping:
                continue
            if any(col == n or col.startswith(n) for n in names):
                mapping[i] = field
                break

    if "measured_at" not in mapping.values():
        log.warning("Cabeceras no reconocidas (%s); se usa el orden documentado",
                    ", ".join(header[:4]))
        mapping = {i: f for i, f in enumerate(POSITIONAL) if f}

    out: dict[datetime, BodyMeasurement] = {}
    for row in table[1:]:
        vals: dict[str, object] = {}
        for i, field in mapping.items():
            if i >= len(row):
                continue
            vals[field] = _when(row[i]) if field == "measured_at" else _num(row[i])
        when = vals.pop("measured_at", None)
        if not when:
            continue
        m = BodyMeasurement(day=when.date(), measured_at=when)
        for field, v in vals.items():
            if v is not None:
                setattr(m, field, int(v) if field == "heart_rate" else v)
        if m.fat_mass_kg is None and m.weight_kg and m.fat_percent:
            m.fat_mass_kg = round(m.weight_kg * m.fat_percent / 100, 2)
        if m.lean_mass_kg is None and m.weight_kg and m.fat_mass_kg:
            m.lean_mass_kg = round(m.weight_kg - m.fat_mass_kg, 2)
        out[when] = m

    log.info("Export de FitDays: %d pesajes (%s -> %s)", len(out),
             min(out).date() if out else "-", max(out).date() if out else "-")
    return [out[k] for k in sorted(out)]
